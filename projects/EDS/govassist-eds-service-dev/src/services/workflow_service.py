from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Type, Union, Set
import tempfile
import pandas as pd
from openpyxl import Workbook, load_workbook
from loguru import logger
from pydantic import BaseModel, ValidationError

from src.repository.workflow_repository import WorkflowRepository
from src.schema.workflow_master_data_schema import (
    WorkflowAccountSchema,
    WorkflowAwardSchema,
    WorkflowCostCenterSchema,
    WorkflowFundSchema,
    WorkflowParentTaskSchema,
    WorkflowProgramSchema,
    WorkflowProjectSchema,
    WorkflowSponsorSchema,
    WorkflowSubTaskSchema,
    WorkflowTransactionSchema,
)
from src.schema.workflow_schema import ExcelSheetName


@dataclass
class MigrationResult:
    """Result object for migration operations"""

    success: bool
    total_records: int
    inserted_records: int
    failed_records: int
    errors: List[str]
    duration_seconds: float
    details: Dict[str, Any]


@dataclass
class ValidationResult:
    """Result object for data validation"""

    valid_records: List[Dict[str, Any]]
    invalid_records: List[Dict[str, Any]]
    validation_errors: List[str]


class ExcelDataMigrationService:
    """Main service for Excel data migration"""

    def __init__(self, repository: WorkflowRepository):
        self.repository = repository

        # Schema mappings for different segments
        self.segment_mappings = {
            "Fund": (WorkflowFundSchema, self.repository.funds),
            "Program-Service Number": (WorkflowProgramSchema, self.repository.programs),
            "Cost Center-Office Number": (
                WorkflowCostCenterSchema,
                self.repository.cost_centers,
            ),
            "Account": (WorkflowAccountSchema, self.repository.accounts),
        }
        # Store the empty structure file path
        self.empty_structure_file_path = None

    def migrate_excel_file(
        self, file_path: Union[str, Path], clear_existing: bool = True
    ):
        """Main method to migrate entire Excel file"""
        start_time = datetime.utcnow()
        try:
            logger.info("Starting Excel migration")

            # Validate file exists
            excel_file = Path(file_path)
            if not excel_file.exists():
                raise FileNotFoundError("Excel file not found")

            # Clear existing data if requested
            if clear_existing:
                deletion_counts = self.repository.clear_all_data()
                logger.info(f"Cleared existing data: {deletion_counts}")

            # Process each sheet in order
            results = {}

            # 1. Chart of Accounts Master Data
            results["coa_master"] = self._process_coa_master_data(excel_file)

            # 2. Award & Project Master Data
            results["award_project"], sub_tasks, project_id_by_number = (
                self._process_award_project_data(excel_file)
            )

            # 3. Summary Balances
            (
                results["summary"],
                sub_tasks,
                project_id_by_number,
                awards_by_number_map,
            ) = self._process_summary_data(excel_file, sub_tasks, project_id_by_number)

            # 4. Transactional Detail Data
            results["transactions"] = self._process_transaction_data(
                excel_file, awards_by_number_map
            )

            # Calculate totals
            total_records = sum(result.total_records for result in results.values())
            inserted_records = sum(
                result.inserted_records for result in results.values()
            )
            failed_records = sum(result.failed_records for result in results.values())
            all_errors = []
            for result in results.values():
                all_errors.extend(result.errors)

            duration = (datetime.utcnow() - start_time).total_seconds()

            logger.info(
                f"Migration completed in {duration:.2f}s. Inserted: {inserted_records}, Failed: {failed_records}"
            )

            return MigrationResult(
                success=failed_records == 0,
                total_records=total_records,
                inserted_records=inserted_records,
                failed_records=failed_records,
                errors=all_errors,
                duration_seconds=duration,
                details=results,
            )

        except Exception as e:
            duration = (datetime.utcnow() - start_time).total_seconds()
            logger.error("Migration failed")
            return MigrationResult(
                success=False,
                total_records=0,
                inserted_records=0,
                failed_records=0,
                errors=[str(e)],
                duration_seconds=duration,
                details={},
            )

    def _process_coa_master_data(self, excel_file: Path) -> MigrationResult:
        """Process Chart of Accounts Master Data sheet"""
        try:
            df = pd.read_excel(
                excel_file,
                sheet_name=ExcelSheetName.COA_MASTER_DATA.value,
                dtype=str,
                header=0,
            )
            df = df.head(1000)

            if "Segment Name" not in df.columns:
                raise ValueError(
                    "'Segment Name' column not found in CoA Master Data sheet"
                )

            # Group data by segment
            grouped_data = defaultdict(list)
            for segment, group_df in df.groupby("Segment Name"):
                records = group_df.to_dict(orient="records")
                grouped_data[segment].extend(records)

            # Process segments in parallel
            results = {}
            with ThreadPoolExecutor(max_workers=1) as executor:
                futures = {}

                for segment_name, records in grouped_data.items():
                    if segment_name not in self.segment_mappings:
                        logger.warning(f"Skipping unknown segment: {segment_name}")
                        continue

                    schema_class, repository = self.segment_mappings[segment_name]
                    future = executor.submit(
                        self._process_segment_data,
                        segment_name,
                        records,
                        schema_class,
                        repository,
                        ExcelSheetName.COA_MASTER_DATA.value,
                    )
                    futures[future] = segment_name

                for future in as_completed(futures):
                    segment_name = futures[future]
                    try:
                        results[segment_name] = future.result()
                    except Exception as e:
                        logger.error(
                            f"Failed to process segment {segment_name}: {str(e)}"
                        )
                        results[segment_name] = MigrationResult(
                            success=False,
                            total_records=0,
                            inserted_records=0,
                            failed_records=0,
                            errors=[str(e)],
                            duration_seconds=0,
                            details={},
                        )

            # Aggregate results
            total_records = sum(r.total_records for r in results.values())
            inserted_records = sum(r.inserted_records for r in results.values())
            failed_records = sum(r.failed_records for r in results.values())
            all_errors = []
            for r in results.values():
                all_errors.extend(r.errors)

            return MigrationResult(
                success=failed_records == 0,
                total_records=total_records,
                inserted_records=inserted_records,
                failed_records=failed_records,
                errors=all_errors,
                duration_seconds=0,
                details=results,
            )

        except Exception as e:
            logger.error("CoA Master Data processing failed")
            return MigrationResult(
                success=False,
                total_records=0,
                inserted_records=0,
                failed_records=0,
                errors=[str(e)],
                duration_seconds=0,
                details={},
            )

    def _process_segment_data(
        self,
        segment_name: str,
        records: List[Dict],
        schema_class: Type[BaseModel],
        repository,
        sheet_name: str,
    ) -> MigrationResult:
        """Process individual segment data with validation"""
        start_time = datetime.utcnow()

        try:
            # Validate data
            validation_result = self._validate_records(records, schema_class)
            # # Store invalid records in a file
            # if validation_result.invalid_records:
            #     self.store_invalid_records(validation_result.invalid_records, sheet_name, self.empty_structure_file_path)

            logger.info(f"valid_records: {len(validation_result.valid_records)}")
            logger.info(f"invalid_records: {len(validation_result.invalid_records)}")
            logger.info(f"validation_errors: {len(validation_result.validation_errors)}")

            if not validation_result.valid_records:
                return MigrationResult(
                    success=False,
                    total_records=len(records),
                    inserted_records=0,
                    failed_records=len(records),
                    errors=validation_result.validation_errors,
                    duration_seconds=(datetime.utcnow() - start_time).total_seconds(),
                    details={"segment": segment_name},
                )
            logger.info(f"valid_records: {len(validation_result.valid_records)}")
            # Convert to model instances
            model_instances = []
            for record in validation_result.valid_records:
                try:
                    schema_obj = schema_class.model_validate(
                        record, from_attributes=False
                    )
                    model_instance = repository.model_class(**schema_obj.model_dump())
                    model_instances.append(model_instance)
                except Exception as e:
                    logger.error(
                        f"Model conversion failed for {segment_name}: {str(e)}"
                    )

            # Bulk insert
            logger.info(f"Model instances: {len(model_instances)}")
            inserted_count = repository.bulk_insert(model_instances)
            logger.info(f"Inserted count: {inserted_count}")
            duration = (datetime.utcnow() - start_time).total_seconds()

            return MigrationResult(
                success=True,
                total_records=len(records),
                inserted_records=inserted_count,
                failed_records=len(validation_result.invalid_records),
                errors=validation_result.validation_errors,
                duration_seconds=duration,
                details={"segment": segment_name},
            )

        except Exception as e:
            duration = (datetime.utcnow() - start_time).total_seconds()
            logger.error(f"Segment {segment_name} processing failed: {str(e)}")
            return MigrationResult(
                success=False,
                total_records=len(records),
                inserted_records=0,
                failed_records=len(records),
                errors=[str(e)],
                duration_seconds=duration,
                details={"segment": segment_name},
            )

    def _process_award_project_data(self, excel_file: Path):
        """Process Award & Project Master Data sheet"""
        try:

            df = pd.read_excel(
                excel_file,
                sheet_name=ExcelSheetName.AWARD_PROJECT_MASTER_DATA.value,
                dtype=str,
                header=0,
            )
            df = df.head(1000)
            df = df.replace({pd.NA: None})

            # TODO: validate award and project master data
            # validation_result = validate_and_handle_invalid_data(df=df)
            # Process hierarchical data
            awards, sponsors, parent_tasks, projects, sub_tasks = (
                self._extract_hierarchical_data(df)
            )

            # Insert awards
            award_result = self._insert_validated_data(
                list(awards.values()), WorkflowAwardSchema, self.repository.awards
            )

            # Insert sponsors
            sponsor_result = self._insert_validated_data(
                list(sponsors.values()), WorkflowSponsorSchema, self.repository.sponsors
            )

            # Resolve project relationships and insert
            self._resolve_project_relationships(projects)
            project_result = self._insert_validated_data(
                list(projects.values()), WorkflowProjectSchema, self.repository.projects
            )
            # Update context with project mappings
            project_id_by_number = self._update_project_context()

            # Insert parent tasks
            self._resolve_parent_task_relationships(parent_tasks)
            parent_task_result = self._insert_validated_data(
                list(parent_tasks.values()),
                WorkflowParentTaskSchema,
                self.repository.parent_tasks,
            )

            # Aggregate results
            total_inserted = (
                award_result.inserted_records
                + sponsor_result.inserted_records
                + project_result.inserted_records
                + parent_task_result.inserted_records
            )

            return (
                MigrationResult(
                    success=True,
                    total_records=len(df),
                    inserted_records=total_inserted,
                    failed_records=0,
                    errors=[],
                    duration_seconds=0,
                    details={
                        "awards": award_result,
                        "sponsors": sponsor_result,
                        "projects": project_result,
                        "parent_tasks": parent_task_result,
                    },
                ),
                sub_tasks,
                project_id_by_number,
            )

        except Exception as e:
            logger.error("Award/Project data processing failed")
            return MigrationResult(
                success=False,
                total_records=0,
                inserted_records=0,
                failed_records=0,
                errors=[str(e)],
                duration_seconds=0,
                details={},
            )

    def _process_summary_data(
        self, excel_file: Path, sub_tasks: Dict, project_id_by_number: Dict
    ) -> MigrationResult:
        """Process Summary Balances sheet"""
        try:
            df = pd.read_excel(
                excel_file,
                sheet_name=ExcelSheetName.SUMMARY_BALANCES.value,
                dtype=str,
                header=0,
            )
            df = df.head(1000)
            df = df.replace({pd.NA: None})
            df.columns = df.columns.str.strip()

            # Enrich sub_tasks with financial data
            sub_tasks = self._enrich_subtasks_with_financial_data(df, sub_tasks)

            # Resolve relationships and insert
            (
                sub_tasks_to_insert,
                sub_tasks,
                project_id_by_number,
                awards_by_number_map,
            ) = self._resolve_subtask_relationships(sub_tasks, project_id_by_number)

            if sub_tasks_to_insert:
                result = self._insert_validated_data(
                    sub_tasks_to_insert,
                    WorkflowSubTaskSchema,
                    self.repository.sub_tasks,
                )
                return result, sub_tasks, project_id_by_number, awards_by_number_map
            else:
                return MigrationResult(
                    success=True,
                    total_records=0,
                    inserted_records=0,
                    failed_records=0,
                    errors=[],
                    duration_seconds=0,
                    details={},
                )

        except Exception as e:
            logger.error("Summary data processing failed")
            return MigrationResult(
                success=False,
                total_records=0,
                inserted_records=0,
                failed_records=0,
                errors=[str(e)],
                duration_seconds=0,
                details={},
            )

    def _process_transaction_data(
        self, excel_file: Path, awards_by_number_map: Dict
    ) -> MigrationResult:
        """Process Transactional Detail Data sheet"""

        try:
            df = pd.read_excel(
                excel_file,
                sheet_name=ExcelSheetName.TRANSACTIONAL_DETAIL_DATA.value,
                dtype=str,
                header=0,
            )
            df = df.head(1000)
            df = df.replace({pd.NA: None})
            df.columns = df.columns.str.strip()

            # Build transaction data
            transaction_data = self._build_transaction_data(df, awards_by_number_map)

            # Insert transaction data
            if transaction_data:
                result = self._insert_validated_data(
                    transaction_data,
                    WorkflowTransactionSchema,
                    self.repository.transactions,
                )
                return result
            else:
                return MigrationResult(
                    success=True,
                    total_records=0,
                    inserted_records=0,
                    failed_records=0,
                    errors=[],
                    duration_seconds=0,
                    details={},
                )

        except Exception as e:
            logger.error("Transaction data processing failed")
            return MigrationResult(
                success=False,
                total_records=0,
                inserted_records=0,
                failed_records=0,
                errors=[str(e)],
                duration_seconds=0,
                details={},
            )

    def _validate_records(
        self, records: List[Dict], schema_class: Type[BaseModel]
    ) -> ValidationResult:
        """Validate records against Pydantic schema"""
        valid_records = []
        invalid_records = []
        validation_errors = []

        for i, record in enumerate(records):
            try:
                schema_class.model_validate(record, from_attributes=False)
                valid_records.append(record)
            except ValidationError as e:
                invalid_records.append(record)
                validation_errors.append(f"Row {i+1}: {str(e)}")

        return ValidationResult(valid_records, invalid_records, validation_errors)

    def _insert_validated_data(
        self, data: List[Dict], schema_class: Type[BaseModel], repository
    ) -> MigrationResult:
        """Insert data with validation"""
        start_time = datetime.utcnow()
        try:
            if not data:
                return MigrationResult(
                    success=True,
                    total_records=0,
                    inserted_records=0,
                    failed_records=0,
                    errors=[],
                    duration_seconds=0,
                    details={},
                )

            # Validate
            validation_result = self._validate_records(data, schema_class)

            # Convert to models
            model_instances = []
            for record in validation_result.valid_records:
                schema_obj = schema_class.model_validate(record, from_attributes=False)
                model_instance = repository.model_class(**schema_obj.model_dump())
                model_instances.append(model_instance)

            # Insert
            inserted_count = repository.bulk_insert(model_instances)
            duration = (datetime.utcnow() - start_time).total_seconds()

            return MigrationResult(
                success=len(validation_result.invalid_records) == 0,
                total_records=len(data),
                inserted_records=inserted_count,
                failed_records=len(validation_result.invalid_records),
                errors=validation_result.validation_errors,
                duration_seconds=duration,
                details={},
            )

        except Exception as e:
            duration = (datetime.utcnow() - start_time).total_seconds()
            return MigrationResult(
                success=False,
                total_records=len(data),
                inserted_records=0,
                failed_records=len(data),
                errors=[str(e)],
                duration_seconds=duration,
                details={},
            )

    def _extract_hierarchical_data(self, df: pd.DataFrame):
        """Extract hierarchical data structures from dataframe"""
        df = df.replace({pd.NA: None})

        # Shared data structures for relationship mapping
        sub_tasks = {}
        awards = {}
        sponsors = {}
        parent_tasks = {}
        projects = {}

        for idx, row in df.iterrows():
            # skip header row.
            if idx == 0:
                continue

            # Award data
            award_key = row["AWARD_NUMBER"]
            if award_key:
                if award_key not in awards:
                    awards[award_key] = {
                        "number": row["AWARD_NUMBER"],
                        "name": row["AWARD_NAME"],
                        "organization": row["AWARD_ORGANIZATION"],
                        "start_date": self._parse_date(row["AWARD_START_DATE"]),
                        "end_date": self._parse_date(row["AWARD_END_DATE"]),
                        "closed_date": self._parse_date(row["AWARD_CLOSE_DATE"]),
                        "status": row["AWARD_STATUS"],
                        "award_type": row["AWARD_TYPE"],
                    }

            # Sponsor data
            sponsor_key = row["SPONSOR_NUMBER"]
            if sponsor_key:
                if sponsor_key not in sponsors:
                    sponsors[sponsor_key] = {
                        "name": row["SPONSOR_NAME"],
                        "number": row["SPONSOR_NUMBER"],
                        "award_number": row["SPONSOR_AWARD_NUMBER"],
                    }

            # Parent task data
            parent_key = (row["PROJECT_NUMBER"], row["PARENT_TASK_NUMBER"])
            if parent_key:
                if parent_key not in parent_tasks:
                    parent_tasks[parent_key] = {
                        "number": row["PARENT_TASK_NUMBER"],
                        "name": row["PARENT_TASK_NAME"],
                        "project_id": row["PROJECT_NUMBER"],
                    }

            # Sub task data
            sub_task_key = (
                row["PROJECT_NUMBER"],
                row["SUB_TASK_NUMBER"],
                row["FUND_NUMBER"],
                row["AWARD_NUMBER"],
            )

            sub_tasks[sub_task_key] = {
                "number": str(float(row["SUB_TASK_NUMBER"])),
                "name": row["SUB_TASK_NAME"],
                "parent_task_id": parent_key,
                "start_date": self._parse_date(row["SUBTASK_START_DATE"]),
                "completion_date": self._parse_date(row["SUBTASK_COMPLETION_DATE"]),
                "award_id": award_key,
                "fund_id": row["FUND_NUMBER"],
            }

            # Project data
            if row["PROJECT_NUMBER"]:
                if row["PROJECT_NUMBER"] not in projects:
                    projects[row["PROJECT_NUMBER"]] = {
                        "owning_agency": row["Owning_Agency"],
                        "principal_investigator": row["PRINCIPAL_INVESTIGATOR"],
                        "number": row["PROJECT_NUMBER"],
                        "name": row["PROJECT_NAME"],
                        "description": row["PROJECT_DESCRIPTION"],
                        "project_type": row["PROJECT_TYPE"],
                        "organization": row["PROJECT_ORGANIZATION"],
                        "master_project_number": row["Master_Project_Number"],
                        "primary_category": row["Primary_Category"],
                        "project_category": row["Project_Category"],
                        "project_classification": row["Project_Classification"],
                        "ward": row["Ward"],
                        "fhwa_improvement_types": row["FHWA_IMPROVEMENT_TYPES"],
                        "fhwa_functional_codes": row["FHWA_FUNCTIONAL_CODE"],
                        "fhwa_capital_outlay_category": row[
                            "FHWA_CAPITAL_OUTLAY_CATEGORY"
                        ],
                        "fhwa_system_code": row["FHWA_SYSTEM_CODE"],
                        "nhs": row["NHS"],
                        "start_date": self._parse_date(row["PROJECT_START_DATE"]),
                        "end_date": self._parse_date(row["PROJECT_END_DATE"]),
                        "status_code": row["PROJECT_STATUS_CODE"],
                        "owner_agency": row["PROJECT_OWNER_AGENCY"],
                        "award_project_burden_schedule_name": row[
                            "Award_Project_Burden_Schedule_Name"
                        ],
                        "iba_project_number": row["IBA_PROJECT_NUMBER"],
                        "burden_rate_multiplier": float(
                            row["Burden_Rate_Multiplier"] or 0
                        ),
                        "burden_schedule_version_start_date": self._parse_date(
                            row["Burden_Schedule_Version_Start_Date"]
                        ),
                        "burden_schedule_version_end_date": self._parse_date(
                            row["Burden_Schedule_Version_End_Date"]
                        ),
                        "burden_schedule_version_name": row[
                            "Burden_Schedule_Version_Name"
                        ],
                        "ind_rate_sch_id": row["IND_RATE_SCH_ID"],
                        "chargeable_flag": row["CHARGEABLE_FLAG"] == "Y",
                        "billable_flag": row["BILLABLE_FLAG"] == "Y",
                        "capitalizable_flag": row["CAPITALIZABLE_FLAG"] == "Y",
                        "cost_center_id": row["COST_CENTER"],
                        "program_id": row["PROGRAM"],
                        "sponsor_id": row["SPONSOR_NUMBER"],
                    }

        return awards, sponsors, parent_tasks, projects, sub_tasks

    def _resolve_project_relationships(self, projects: Dict):
        """Resolve foreign key relationships for projects"""
        # Get ID mappings
        program_mapping = self.repository.programs.get_id_mapping()
        cost_center_mapping = self.repository.cost_centers.get_id_mapping()
        sponsor_mapping = self.repository.sponsors.get_id_mapping()

        for project_data in projects.values():
            project_data["program_id"] = program_mapping.get(
                project_data.get("program_id")
            )
            project_data["cost_center_id"] = cost_center_mapping.get(
                project_data.get("cost_center_id")
            )
            project_data["sponsor_id"] = sponsor_mapping.get(
                project_data.get("sponsor_id")
            )

    def _update_project_context(self):
        """Update context with project ID mappings"""
        project_id_by_number = {}
        project_mapping = self.repository.projects.get_id_mapping()
        for number, project_id in project_mapping.items():
            project_id_by_number[project_id] = number
        return project_id_by_number

    def _resolve_parent_task_relationships(self, parent_tasks: Dict):
        """Resolve parent task relationships"""
        project_mapping = self.repository.projects.get_id_mapping()
        for task_data in parent_tasks.values():
            task_data["project_id"] = project_mapping.get(task_data.get("project_id"))

    def _enrich_subtasks_with_financial_data(self, df: pd.DataFrame, sub_tasks: Dict):
        """Enrich sub tasks with financial summary data"""

        for idx, row in df.iterrows():
            if idx == 0:
                continue

            # Sub task data
            task_key = (
                str(row.get("PROJECT", "")).strip(),
                str(row.get("TASK NUMBER", "")).strip(),
                str(row.get("FUND", "")).strip(),
                str(row.get("AWARD", "")).strip(),
            )

            task = sub_tasks.get(task_key)
            if task:
                sub_tasks.get(task_key).update(
                    {
                        "award_funding_amount": float(
                            row.get("AWARD_FUNDING_AMOUNT") or 0
                        ),
                        "png_lifetime_budget": float(
                            row.get("PNG_LIFETIME_BUDGET") or 0
                        ),
                        "png_lifetime_allotment": float(
                            row.get("PNG_LIFETIME_ALLOTMENT") or 0
                        ),
                        "commitment": float(row.get("COMMITMENT") or 0),
                        "obligation": float(row.get("OBLIGATION") or 0),
                        "expenditure": float(row.get("EXPENDITURE") or 0),
                        "receivables": float(row.get("RECEIVABLES") or 0),
                        "revenue": float(row.get("REVENUE") or 0),
                    }
                )

        return sub_tasks

    def _resolve_subtask_relationships(
        self, sub_tasks: Dict, project_id_by_number: Dict
    ):
        """Resolve sub task relationships"""

        # Get ID mappings
        awards_by_number_map = {}
        fund_mapping = self.repository.funds.get_id_mapping()
        award_mapping = self.repository.awards.get_id_mapping()
        awards_by_number_map.update(award_mapping)

        # Get parent task mapping
        parent_task_mapping = self.repository.get_parent_task_mapping(
            project_id_by_number
        )

        # Insert sub tasks
        sub_tasks_to_insert = []
        for _, sub_task_data in sub_tasks.items():
            parent_task_id = parent_task_mapping.get(sub_task_data["parent_task_id"])
            if parent_task_id is not None:
                fund_id = fund_mapping.get(sub_task_data["fund_id"])
                award_id = awards_by_number_map.get(sub_task_data["award_id"])

                # TODO: Uncomment this when we have a way to handle non-existent fund_id and award_id
                # if fund_id is None:
                #     logger.error(f"non-existent fund_id: {sub_task_data}")
                # if award_id is None:
                #     logger.error(f"non-existent award_id: {sub_task_data}")

                sub_task_data["parent_task_id"] = parent_task_id
                sub_task_data["fund_id"] = fund_id
                sub_task_data["award_id"] = award_id
                sub_tasks_to_insert.append(sub_task_data)

        return (
            sub_tasks_to_insert,
            sub_tasks,
            project_id_by_number,
            awards_by_number_map,
        )

    def _build_transaction_data(
        self, df: pd.DataFrame, awards_by_number_map: Dict
    ) -> List[Dict]:
        """Build transaction data with relationships"""

        # Get sub task mapping
        sub_task_by_project_award_map = (
            self.repository._get_sub_task_by_project_award_map()
        )

        not_found_sub_task_keys = []
        transaction_data = []

        for idx, row in df.iterrows():
            if idx == 0:
                continue

            # Sub task data
            sub_task_key = (
                self._safe_strip(row["Project Number"]),
                str(float(self._safe_strip(row["Subtask Number"]))),
                self._safe_strip(row["Award Number"]),
            )
            sub_task_id = sub_task_by_project_award_map.get(sub_task_key)

            # Insert transaction data
            if sub_task_id is not None:
                data = {
                    "sub_task_id": sub_task_id,
                    "award_id": awards_by_number_map.get(
                        self._safe_strip(row["Award Number"])
                    ),
                    "transaction_number": self._safe_strip(row["Transaction Number"]),
                    "transaction_source": self._safe_strip(row["Transaction Source"]),
                    "expenditure_type": self._safe_strip(row["Expenditure Type"]),
                    "expenditure_category": self._safe_strip(
                        row["Expenditure Category"]
                    ),
                    "expenditure_organization": self._safe_strip(
                        row["Expenditure Organization"]
                    ),
                    "expenditure_item_date": self._parse_date(
                        row["Expenditure Item Date"]
                    ),
                    "accounting_period": self._safe_strip(row["Accounting Period"]),
                    "unit_of_measure": self._safe_strip(row["Unit of Measure"]),
                    "incurred_by_person": self._safe_strip(row["Incurred By Person"]),
                    "person_number": self._safe_strip(row["Person Number"]),
                    "position_number": self._safe_strip(row["Position Number"]),
                    "vendor_name": self._safe_strip(row["Vendor Name"]),
                    "po_number": self._safe_strip(row["PO Number"]),
                    "po_line_number": self._safe_strip(row["PO Line Number"]),
                    "ap_invoice_number": self._safe_strip(row["AP Invoice Number"]),
                    "ap_invoice_line_number": self._safe_strip(
                        row["AP Invoice Line Number"]
                    ),
                    "dist_line_num": self._safe_strip(row["Dist Line Num"]),
                    "invoice_date": self._parse_date(row["Invoice Date"]),
                    "check_number": self._safe_strip(row["Check Number"]),
                    "check_date": self._parse_date(row["Check Date"]),
                    "expenditure_batch": self._safe_strip(row["Expenditure Batch"]),
                    "expenditure_comment": self._safe_strip(row["Expenditure Comment"]),
                    "orig_transaction_reference": self._safe_strip(
                        row["Orig Transaction Reference"]
                    ),
                    "capitalizable_flag": self._safe_strip(row["Capitalizable Flag"])
                    == "Y",
                    "billable_flag": self._safe_strip(row["Billable Flag"]) == "Y",
                    "bill_hold_flag": self._safe_strip(row["Bill Hold Flag"]) == "Y",
                    "revenue_status": self._safe_strip(row["Revenue status"]),
                    "transaction_ar_invoice_status": self._safe_strip(
                        row["Transaction AR Invoice Status"]
                    ),
                    "servicedate_from": self._parse_date(row["ServiceDate From"]),
                    "servicedate_to": self._parse_date(row["ServiceDate To"]),
                    "gl_batch_name": self._safe_strip(row["GL Batch Name"]),
                    "quantity": float(row["Quantity"] or 0),
                    "transaction_amount": float(row["Transaction Amount"] or 0),
                    "burdened_amount": float(row["Burdened Amount"] or 0),
                    "rate": float(row["Rate"] or 0),
                }
                transaction_data.append(data)
            else:
                not_found_sub_task_keys.append(sub_task_key)

        if not_found_sub_task_keys:
            logger.error(
                f"Not found {len(not_found_sub_task_keys)} sub task keys: {not_found_sub_task_keys}"
            )

        return transaction_data

    def _safe_strip(self, value):
        """Safely strip string values"""
        if value is None or pd.isna(value):
            return None
        return str(value).strip()

    def _parse_date(self, value):
        """Parse date value safely"""
        if pd.isna(value) or value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        try:
            return pd.to_datetime(value).date()
        except Exception:
            logger.warning(f"Error parsing date: {value}")
            return None

    
    def store_invalid_records(self, invalid_data: List[Dict], sheet_name: str, excel_file_path: str) -> None:
        # Convert data to DataFrame
        df_to_append = pd.DataFrame(invalid_data)
        df_to_append = df_to_append.where(pd.notna(df_to_append), None)

        # Load workbook to get current row count
        book = load_workbook(excel_file_path)
        if sheet_name not in book.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")

        # Determine the starting row (after last row with data)
        start_row = book[sheet_name].max_row

        # Append using modern pandas API (2.1+)
        with pd.ExcelWriter(
            excel_file_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='overlay'
        ) as writer:
            df_to_append.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=start_row
            )
    def create_empty_excel_with_same_structure(self, excel_file: Path) -> str:
        try:
            # Read the original Excel file to get sheet names and headers
            excel = pd.ExcelFile(excel_file)
            sheet_names = excel.sheet_names
            
            # Create a temporary file
            temp_file = tempfile.NamedTemporaryFile(
                suffix=".xlsx", 
                prefix="empty_structure_", 
                delete=False
            )
            temp_file.close()
            
            # Create a new workbook
            workbook = Workbook()
            
            # Remove the default sheet
            if workbook.active:
                workbook.remove(workbook.active)
            
            # Create sheets with the same names and headers as the original file
            for sheet_name in sheet_names:
                # Excel sheet names are limited to 31 characters
                safe_sheet_name = sheet_name[:31]
                worksheet = workbook.create_sheet(title=safe_sheet_name)
                
                # Read the original sheet to get headers
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=0)
                    headers = df.columns.tolist()
                    
                    # Add headers to the new sheet
                    worksheet.append(headers)
                    
                    # Style the header row (make it bold)
                    for cell in worksheet[1]:
                        cell.font = cell.font.copy(bold=True)
                        
                except Exception as e:
                    logger.warning(f"Could not read headers from sheet '{sheet_name}': {e}")
                    # Add a default header if we can't read the original
                    worksheet.append(["Column 1"])
            
            # Save the workbook
            workbook.save(temp_file.name)
            workbook.close()
            
            # Store the file path
            self.empty_structure_file_path = temp_file.name
            
            logger.info(f"Created empty Excel file with same structure: {temp_file.name}")
            return temp_file.name
            
        except Exception as e:
            logger.error(f"Failed to create empty Excel file with same structure: {e}")
            return None
