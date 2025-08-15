import copy
import os
from datetime import datetime
from io import BytesIO
from typing import Any, Callable, List, Optional

from fastapi.responses import StreamingResponse
from loguru import logger
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload, make_transient, selectinload

from src.core.config import configs
from src.core.exceptions import NotFoundError
from src.model.budget_info_model import BudgetInfo
from src.model.elasticsearch_models import ParModel
from src.model.par_activity_model import ParActivity
from src.model.par_award_association_model import ParAwardAssociation
from src.model.par_budget_analysis_model import ParBudgetAnalysis
from src.model.par_model import Par
from src.model.project_detail_model import ProjectDetails
from src.model.reviewed_par_model import ReviewedPar
from src.repository.par_repository import ParRepository
from src.schema.elastic_search_schema import ParES
from src.schema.key_metrics_schema import KeyMetrics
from src.schema.par_activity_schema import ParActivityCreate
from src.schema.par_meta_schema import ParMeta
from src.schema.par_schema import ParCreate, ParFind, ParUpdate
from src.services.base_service import BaseService
from src.services.par_activity_service import ParActivityService
from src.util.budget_analyst_function import group_and_sum_change_amount
from src.util.export_excel_function import copy_row_styles
from src.util.par import get_latest_par_status_map
from src.util.query_builder import dict_to_sqlalchemy_filter_options


class ParService(BaseService):
    def __init__(
        self,
        session_factory: Callable[..., Session],
        par_activity_service: ParActivityService,
    ):
        super().__init__(ParRepository(session_factory))
        self._session_factory = session_factory
        self._par_activity_service = par_activity_service
        self._par_model = ParModel()

    async def add(self, par: ParCreate):
        par_data = super().add(par)

        # Create PAR document in Elasticsearch first
        es_data = self._par_model._get_clean_model_data(par_data)

        try:
            # Then create PAR activity
            par_activity = await self._par_activity_service.add(
                ParActivityCreate(
                    par_id=par_data.id,
                    user="Quality Assurance",
                    date=datetime.now(),
                    activity="PAR Created",
                    status="DRAFT",
                    comments="Created PAR",
                ),
                update_elasticsearch=False,
            )
            activity_data = self._par_model._get_clean_model_data(par_activity)
            es_data["par_activities"] = [activity_data]
            await self._par_model.create_par(par_data.id, es_data, refresh=True)
        except Exception as e:
            logger.error(f"Error updating PAR activity in Elasticsearch: {str(e)}")
            # Don't raise the error since the PAR was created successfully
            # The activity can be updated later

        return par_data

    async def patch(self, par_id: int, par: ParUpdate):
        db_par = super().patch(par_id, par)
        es_data = self._par_model._get_clean_model_data(db_par)
        # NOTE: For now we are not updating the project_details_id in the par so here also we are not updating the project_details in the elasticsearch.
        await self._par_model.update_par(par_id, es_data)
        return db_par

    def clone_model(self, obj, exclude: tuple = (), overrides: dict = {}):

        # Get the data from the object's columns
        data = {
            col.name: getattr(obj, col.name)
            for col in obj.__table__.columns
            if col.name not in exclude
        }
        data.update(overrides)

        # Create a new instance that's completely detached
        new_obj = obj.__class__(**data)

        # Make the object transient (not associated with any session)
        make_transient(new_obj)

        return new_obj

    async def par_clone(self, par_id: int, request_type: Optional[str] = None):
        # First, load the existing PAR data using the repository method which handles sessions properly

        new_par_id = None
        # Now create the clone in a fresh session
        with self._session_factory() as session:
            existing_par = (
                session.query(Par)
                .options(
                    selectinload(Par.project_details).selectinload(
                        ProjectDetails.par_award_associations
                    ),
                    selectinload(Par.par_budget_analysis).selectinload(
                        ParBudgetAnalysis.additional_fund
                    ),
                    selectinload(Par.par_budget_analysis).selectinload(
                        ParBudgetAnalysis.par_budget_analysis_fund
                    ),
                    selectinload(Par.budget_info).selectinload(BudgetInfo.budget_items),
                )
                .filter(Par.id == par_id)
                .first()
            )
            if not existing_par:
                raise ValueError(f"PAR with ID {par_id} not found")
            # Create new PAR instance
            new_par = self.clone_model(
                existing_par, exclude=("id", "uuid", "created_at", "updated_at")
            )
            if request_type:
                new_par.request_type = request_type

            # Clear relationships to avoid session conflicts
            new_par.project_details = None
            new_par.par_budget_analysis = []
            new_par.budget_info = []
            new_par.par_activities = []

            # Add and flush to get the ID
            session.add(new_par)
            session.flush()
            new_par_id = new_par.id

            # Clone project details if they exist
            new_project_detail = None
            if existing_par.project_details:
                new_project_detail = self.clone_model(
                    existing_par.project_details,
                    exclude=("id", "created_at", "updated_at", "uuid"),
                )
                session.add(new_project_detail)
                session.flush()
                new_par.project_details_id = new_project_detail.id

                # Clone par award associations
                if (
                    hasattr(existing_par.project_details, "par_award_associations")
                    and existing_par.project_details.par_award_associations
                ):
                    for (
                        par_award_association
                    ) in existing_par.project_details.par_award_associations:
                        new_association = self.clone_model(
                            par_award_association,
                            exclude=("id", "created_at", "updated_at", "uuid"),
                            overrides={"project_details_id": new_project_detail.id},
                        )
                        session.add(new_association)

            # Clone par_budget_analysis records if they exist
            if existing_par.par_budget_analysis:
                for analysis in existing_par.par_budget_analysis:
                    new_analysis = self.clone_model(
                        analysis,
                        exclude=(
                            "id",
                            "created_at",
                            "updated_at",
                            "uuid",
                            "par_id",
                            "project_details_id",
                        ),
                        overrides={
                            "par_id": new_par.id,
                            "project_details_id": (
                                new_project_detail.id if new_project_detail else None
                            ),
                        },
                    )
                    session.add(new_analysis)
                    session.flush()
                    for additional_fund in analysis.additional_fund:
                        new_additional_fund = self.clone_model(
                            additional_fund,
                            exclude=(
                                "id",
                                "created_at",
                                "updated_at",
                                "uuid",
                                "par_budget_analysis_id",
                            ),
                            overrides={"par_budget_analysis_id": new_analysis.id},
                        )
                        session.add(new_additional_fund)
                    for par_budget_analysis_fund in analysis.par_budget_analysis_fund:
                        new_par_budget_analysis_fund = self.clone_model(
                            par_budget_analysis_fund,
                            exclude=(
                                "id",
                                "created_at",
                                "updated_at",
                                "uuid",
                                "par_budget_analysis_id",
                            ),
                            overrides={"par_budget_id": new_analysis.id},
                        )
                        session.add(new_par_budget_analysis_fund)

            # Clone budget info and items
            if existing_par.budget_info:
                for budget_info in existing_par.budget_info:
                    new_budget_info = self.clone_model(
                        budget_info,
                        exclude=(
                            "id",
                            "created_at",
                            "updated_at",
                            "uuid",
                            "par_id",
                            "budget_items",
                        ),
                        overrides={
                            "par_id": new_par.id,
                        },
                    )
                    session.add(new_budget_info)
                    session.flush()

                    # Clone budget items
                    if (
                        hasattr(budget_info, "budget_items")
                        and budget_info.budget_items
                    ):
                        for budget_item in budget_info.budget_items:
                            new_budget_item = self.clone_model(
                                budget_item,
                                exclude=(
                                    "id",
                                    "created_at",
                                    "updated_at",
                                    "uuid",
                                    "budget_info_id",
                                ),
                                overrides={"budget_info_id": new_budget_info.id},
                            )
                            session.add(new_budget_item)

            session.commit()

            # Create PAR activity
            activity_data = ParActivityCreate(
                par_id=new_par_id,
                user="System",
                date=datetime.now(),
                activity="PAR Cloned",
                status="DRAFT",
                comments=f"Cloned from PAR ID: {par_id}",
            )
            await self._par_activity_service.add(
                activity_data, update_elasticsearch=False
            )

            # Load the complete PAR data with all relationships for Elasticsearch
            with self._session_factory() as es_session:
                complete_par = (
                    es_session.query(Par)
                    .options(
                        selectinload(Par.project_details)
                        .selectinload(ProjectDetails.par_award_associations)
                        .selectinload(ParAwardAssociation.award_type),
                        selectinload(Par.budget_info),
                        selectinload(Par.par_activities),
                    )
                    .filter(Par.id == new_par_id)
                    .first()
                )

                if complete_par:
                    try:
                        # Use Pydantic from_orm to convert SQLAlchemy object to ES schema
                        par_es_model = ParES.from_orm(complete_par)
                        # Add the new activity to par_activities
                        par_dict = par_es_model.dict()
                        # Create Elasticsearch document
                        await self._par_model.create_par(
                            new_par_id, par_dict, refresh=True
                        )
                        logger.info(
                            f"Successfully created Elasticsearch document for cloned PAR {new_par_id}"
                        )

                    except Exception as e:
                        logger.error(
                            f"Failed to create Elasticsearch document for cloned PAR {new_par_id}: {str(e)}"
                        )
                        # Don't fail the whole operation if ES indexing fails

            return "Clone PAR successfully", new_par_id

    async def get_par_meta(self, par_id: int):
        par_data = self.get_by_id(par_id)
        latest_status = self._par_activity_service.fetch_par_latest_activity_status(
            par_id
        )

        par_meta = ParMeta(
            epar_name=par_data.epar_name,
            request_type=par_data.request_type,
            total_project_budget=par_data.total_project_budget,
            funding_source=(
                par_data.project_details.funding_source
                if par_data.project_details
                else None
            ),
            project_name=(
                par_data.project_details.project_name
                if par_data.project_details
                else None
            ),
            icrs_exempt=(
                par_data.project_details.icrs_exempt
                if par_data.project_details
                else None
            ),
            icrs_rate=(
                par_data.project_details.icrs_rate if par_data.project_details else None
            ),
            project_details_id=(
                par_data.project_details.id if par_data.project_details else None
            ),
            current_status=latest_status,
            user=par_data.par_activities[-1].user if par_data.par_activities else None,
            updated_date=(
                par_data.par_activities[-1].date.date()
                if par_data.par_activities
                else None
            ),
        )

        return par_meta

    async def get_par_excel(self, par_id: int):
        db_par = self.get_by_id(par_id)

        if db_par is None:
            raise NotFoundError(detail="PAR not found")

        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        FILE_PATH = os.path.join(BASE_DIR, "..", "Sample_Budget.xlsx")
        MIDDLE_SHEET_NAME = "COMBO-PAR (budget changes)"

        if not os.path.exists(FILE_PATH):
            raise NotFoundError(detail="Template file not found")

        wb = load_workbook(FILE_PATH)
        if MIDDLE_SHEET_NAME not in wb.sheetnames:
            raise NotFoundError(detail="Sheet not found")

        original_sheet = wb[MIDDLE_SHEET_NAME]

        # Create a new workbook and copy the sheet
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = MIDDLE_SHEET_NAME

        total_changes = group_and_sum_change_amount(db_par.budget_info)

        # Copy data and formatting
        for row in original_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(
                    row=cell.row, column=cell.column, value=cell.value
                )
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)

        # Copy merged cells
        for merged_range in original_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))

        # Copy column widths and adjust specific ones
        for col_letter, col_dim in original_sheet.column_dimensions.items():
            width = col_dim.width
            if col_letter in ["D", "E"]:
                width *= 0.75  # Reduce width of D and E
            new_sheet.column_dimensions[col_letter].width = width

        new_sheet.column_dimensions["F"].width = 30  # Set width of F
        new_sheet.column_dimensions["G"].width = 30
        new_sheet.column_dimensions["H"].width = 30
        new_sheet.column_dimensions["I"].width = 43

        # Copy row heights
        for row_idx, row_dim in original_sheet.row_dimensions.items():
            if row_dim.height:
                new_sheet.row_dimensions[row_idx].height = row_dim.height

        # Add dynamic data
        new_sheet.cell(
            row=5, column=4, value=db_par.request_type if db_par.request_type else ""
        )
        new_sheet.cell(row=6, column=4, value="=TODAY()")
        new_sheet.cell(
            row=7,
            column=4,
            value=(
                db_par.project_details.project_manager
                if db_par.project_details and db_par.project_details.project_manager
                else ""
            ),
        )
        new_sheet.cell(row=8, column=4, value="")  # email
        new_sheet.cell(row=9, column=4, value="")  # phone
        new_sheet.cell(row=10, column=4, value="")  # RAD Approval
        new_sheet.cell(
            row=13,
            column=4,
            value=(
                db_par.project_details.project_name
                if db_par.project_details and db_par.project_details.project_name
                else ""
            ),
        )
        new_sheet.cell(row=14, column=4, value="")  # Project Name
        new_sheet.cell(
            row=15,
            column=4,
            value=(
                db_par.project_details.project_number
                if db_par.project_details and db_par.project_details.project_number
                else ""
            ),
        )  # Difs Project Number
        new_sheet.cell(
            row=16,
            column=4,
            value=(
                db_par.project_details.fhwa_soar_project_no.soar_project_no
                if db_par.project_details
                and db_par.project_details.fhwa_soar_project_no
                and db_par.project_details.fhwa_soar_project_no.soar_project_no
                else ""
            ),
        )  # Soar Project Number
        new_sheet.cell(
            row=17,
            column=4,
            value=(
                db_par.project_details.cost_center.cost_center_name
                if db_par.project_details
                and db_par.project_details.cost_center
                and db_par.project_details.cost_center.cost_center_name
                else ""
            ),
        )  # Difs Cost Center
        new_sheet.cell(
            row=18,
            column=4,
            value=(
                db_par.project_details.program_code
                if db_par.project_details and db_par.project_details.program_code
                else ""
            ),
        )  # Difs Program Code
        new_sheet.cell(
            row=19,
            column=4,
            value=(
                "Exempt"
                if getattr(db_par.project_details, "icrs_exempt", False)
                else "Not Exempt"
            ),
        )  # IDCR
        new_sheet.cell(
            row=20,
            column=4,
            value=(
                db_par.project_details.master_project.master_project_number
                if db_par.project_details
                and db_par.project_details.master_project
                and db_par.project_details.master_project.master_project_number
                else ""
            ),
        )  # Master Project Number
        new_sheet.cell(row=21, column=4, value="")  # PoDI or state Assumed
        new_sheet.cell(row=22, column=4, value="")  # Program
        new_sheet.cell(
            row=23,
            column=4,
            value=(
                db_par.project_details.fhwa_soar_project_no.project_number
                if db_par.project_details
                and db_par.project_details.fhwa_soar_project_no
                and db_par.project_details.fhwa_soar_project_no.project_number
                else ""
            ),
        )  # Federal-Aid Project Number
        new_sheet.cell(
            row=24,
            column=4,
            value=(
                db_par.project_details.fhwa_soar_project_no.stip_reference
                if db_par.project_details
                and db_par.project_details.fhwa_soar_project_no
                and db_par.project_details.fhwa_soar_project_no.stip_reference
                else ""
            ),
        )  # Stip number

        funding_source_map = {
            "federal_grant": "Federal",
            "local": "Local",
        }
        funding_key = getattr(db_par.project_details, "funding_source", "")
        new_sheet.cell(
            row=25,
            column=4,
            value=(funding_source_map.get(funding_key, "")),
        )  # Funding source
        new_sheet.cell(
            row=26,
            column=4,
            value=(
                db_par.project_details.contract_number
                if db_par.project_details and db_par.project_details.contract_number
                else ""
            ),
        )  # Contract Number
        new_sheet.cell(row=27, column=4, value="")  # IDIQ Contact -
        new_sheet.cell(row=28, column=4, value="")  # Contract Start Date
        new_sheet.cell(
            row=29,
            column=4,
            value=(
                db_par.project_details.current_project_end_date
                if db_par.project_details
                and db_par.project_details.current_project_end_date
                else ""
            ),
        )  # Project End Date

        new_sheet.cell(
            row=5, column=7, value=db_par.justification if db_par.justification else ""
        )  # Justification

        new_sheet.cell(row=7, column=7, value="")  # Administration -
        new_sheet.cell(row=8, column=7, value="")  # Divion -
        new_sheet.cell(row=9, column=7, value="")  # Branch -
        new_sheet.cell(row=13, column=7, value=db_par.description)  # Description
        new_sheet.cell(row=16, column=7, value="")  # DIFS Primary Category -
        new_sheet.cell(row=17, column=7, value="")  # DIFS Project Category -
        new_sheet.cell(row=18, column=7, value="")  # DIFS Project Classification
        new_sheet.cell(
            row=19,
            column=7,
            value=(
                db_par.project_details.asset_type
                if db_par.project_details and db_par.project_details.asset_type
                else ""
            ),
        )  # Asset Type
        new_sheet.cell(
            row=20,
            column=7,
            value=(
                db_par.project_details.bridge_number
                if db_par.project_details
                and db_par.project_details.fhwa_soar_project_no
                and db_par.project_details.fhwa_soar_project_no.stip_reference
                else ""
            ),
        )  # Bridge Number (NBIS #)
        new_sheet.cell(
            row=21,
            column=7,
            value=(
                db_par.project_details.improvement_type
                if db_par.project_details and db_par.project_details.improvement_type
                else ""
            ),
        )  # Improvement Type
        new_sheet.cell(
            row=22,
            column=7,
            value=(
                db_par.project_details.project_location.location
                if db_par.project_details
                and db_par.project_details.project_location
                and db_par.project_details.project_location.location
                else ""
            ),
        )  # Ward
        new_sheet.cell(
            row=23,
            column=7,
            value=(
                db_par.project_details.gis_route_id
                if db_par.project_details and db_par.project_details.gis_route_id
                else ""
            ),
        )  # GIS Route ID
        new_sheet.cell(
            row=24,
            column=7,
            value=(
                db_par.project_details.beginning_point
                if db_par.project_details and db_par.project_details.beginning_point
                else ""
            ),
        )  # Beginning Point
        new_sheet.cell(
            row=25,
            column=7,
            value=(
                db_par.project_details.end_point
                if db_par.project_details and db_par.project_details.end_point
                else ""
            ),
        )  # End Point
        new_sheet.cell(row=26, column=7, value="")  # NEPA Type (Obligations Only)
        new_sheet.cell(row=27, column=7, value="")  # NEPA Signed Date (Obs Only)
        new_sheet.cell(row=28, column=7, value="")  # FMIS Authorization Type
        new_sheet.cell(row=29, column=7, value="")  # FMIS Authorization Date
        new_sheet.cell(row=30, column=7, value="")  # NHS Network - ON/OFF

        feasibility_start_row = 34
        design_start_row = rights_of_way_start_row = project_management_start_row = 4
        construction_start_row = equipment_start_row = final_row = 4

        is_exempt = (
            db_par.project_details.icrs_exempt if db_par.project_details else False
        )
        column_mapping = {"task_name": 2, "index": 3, "account": 4, "change_amount": 5}

        def process_task(task_name, start_row, next_section_start_row):
            index = 0
            exempt_values = []
            row_level = start_row

            for task in db_par.budget_info:
                if task.task_name.lower() == task_name.lower():
                    budget_items = getattr(task, "budget_items", [])
                    for item in budget_items:
                        new_sheet.cell(
                            row=row_level, column=column_mapping["index"], value=index
                        )
                        new_sheet.cell(
                            row=row_level,
                            column=column_mapping["account"],
                            value=item.account,
                        )
                        new_sheet.cell(
                            row=row_level,
                            column=column_mapping["change_amount"],
                            value=item.change_amount,
                        )

                        index += 1
                        row_level += 1

                        new_sheet.insert_rows(row_level + 1)
                        copy_row_styles(new_sheet, row_level, row_level + 1)
                        next_section_start_row += 1
                        exempt_values.append(item.change_amount)

            total = sum(value for value in exempt_values if value is not None)
            if not is_exempt and exempt_values:
                new_sheet.cell(
                    row=next_section_start_row + start_row - 2,
                    column=column_mapping["change_amount"],
                    value=total * 0.105,
                )
            new_sheet.cell(
                row=next_section_start_row + start_row - 1,
                column=column_mapping["change_amount"],
                value=total,
            )

            return next_section_start_row + start_row

        design_start_row = process_task(
            "Feasibility Studies", feasibility_start_row, design_start_row
        )
        rights_of_way_start_row = process_task(
            "Design", design_start_row, rights_of_way_start_row
        )
        project_management_start_row = process_task(
            "Rights of Way", rights_of_way_start_row, project_management_start_row
        )
        construction_start_row = process_task(
            "Project Management", project_management_start_row, construction_start_row
        )

        equipment_start_row = process_task(
            "Construction", construction_start_row, equipment_start_row
        )

        final_row = process_task("Equipment", equipment_start_row, final_row)

        categories = [
            "Project Management",
            "Construction",
            "Feasibility Studies",
            "Design",
            "Equipment",
            "Rights of Way",
        ]

        totals = [total_changes.get(category, 0) for category in categories]
        grand_total = sum(totals)
        total_row = final_row
        cell = new_sheet.cell(
            row=total_row, column=column_mapping["change_amount"], value=grand_total
        )
        new_sheet.row_dimensions[total_row].height = 30

        # Save the new file
        # Save the workbook to an in-memory BytesIO stream
        output_stream = BytesIO()
        new_wb.save(output_stream)
        output_stream.seek(0)  # Move to the beginning

        filename = f"PAR_{db_par.epar_name}_{db_par.id}.xlsx"

        return StreamingResponse(
            output_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    def get_par_activity(self, par_id: int):
        par_data = self.get_by_id(par_id)
        if par_data.par_activities:
            return [activity for activity in par_data.par_activities]
        else:
            raise NotFoundError(detail=f"No activities found for this PAR {par_id}")

    def get_key_metrics(self):
        with self._session_factory() as session:
            # Get total count of PARs
            total_pars = session.query(func.count(Par.id)).scalar()

            # Get total amount
            total_amount = session.query(
                func.coalesce(func.sum(Par.total_project_budget), 0.0)
            ).scalar()

            # Get count of PARs with DRAFT status in their latest activity
            # Using a subquery to get the latest activity for each PAR
            latest_activities = session.query(
                ParActivity.par_id,
                ParActivity.status,
                func.row_number()
                .over(partition_by=ParActivity.par_id, order_by=ParActivity.date.desc())
                .label("rn"),
            ).subquery()

            draft_pars = (
                session.query(func.count(Par.id))
                .join(latest_activities, Par.id == latest_activities.c.par_id)
                .filter(
                    latest_activities.c.rn == 1, latest_activities.c.status == "DRAFT"
                )
                .scalar()
            ) or 0

        return KeyMetrics(
            total_pars=total_pars,
            pending_pars=draft_pars,
            active_wards=3,
            total_amount=total_amount,
        )

    def get_recently_reviewed_pars(
        self,
        user_id: str,
        find: Optional[ParFind] = None,
        searchable_fields: Optional[List[str]] = None,
    ) -> dict:
        """
        Get recently reviewed PARs for a specific user with support for searching, filtering, and sorting.

        Args:
            user_id: The ID of the user
            find: Search and filter criteria (ParFind schema)
            searchable_fields: List of fields to search in

        Returns:
            Dictionary containing the list of recently reviewed PARs and search options
        """
        if find is None:
            find = ParFind()

        with self._session_factory() as session:
            # Get the schema as dict for processing
            schema_as_dict: dict = find.dict(exclude_none=True)

            # Handle searchable fields
            if isinstance(searchable_fields, str):
                searchable_fields = [searchable_fields]
            if searchable_fields is None and schema_as_dict.get("search"):
                searchable_fields = ["epar_name"]

            # Get pagination and ordering parameters
            ordering: str = schema_as_dict.get("ordering", configs.ORDERING)
            order_query = (
                getattr(Par, ordering[1:]).desc()
                if ordering.startswith("-")
                else getattr(Par, ordering).asc()
            )
            page = schema_as_dict.get("page", configs.PAGE)
            page_size = schema_as_dict.get("page_size", configs.PAGE_SIZE)
            search_term = schema_as_dict.get("search")

            # Base query with join to ReviewedPar table and eager loading
            query = (
                session.query(Par)
                .join(ReviewedPar, ReviewedPar.par_id == Par.id)
                .filter(ReviewedPar.user_id == user_id)
            )

            # Add eager loading for related data
            for relation_path in getattr(Par, "eagers", []):
                path_parts = relation_path.split(".")
                current_class = Par
                current_attr = getattr(current_class, path_parts[0])
                loader = joinedload(current_attr)

                for part in path_parts[1:]:
                    current_class = current_attr.property.mapper.class_
                    current_attr = getattr(current_class, part)
                    loader = loader.joinedload(current_attr)

                query = query.options(loader)

            # Apply filters (excluding search and pagination parameters)
            filter_dict = find.dict(exclude_none=True)
            excluded_keys = ["search", "page", "page_size", "ordering"]
            for key in excluded_keys:
                if key in filter_dict:
                    del filter_dict[key]

            if filter_dict:
                filter_options = dict_to_sqlalchemy_filter_options(Par, filter_dict)
                query = query.filter(filter_options)

            # Apply search filters
            if searchable_fields and search_term:
                search_filters = []
                for field in searchable_fields:
                    if hasattr(Par, field):
                        search_filters.append(
                            getattr(Par, field).ilike(f"%{search_term}%")
                        )
                if search_filters:
                    query = query.filter(or_(*search_filters))

            # Order by ReviewedPar.updated_at desc first (for recency), then by user-specified ordering
            query = query.order_by(ReviewedPar.updated_at.desc(), order_query)

            # Get total count
            total_count = query.count()

            # Apply pagination
            if page_size == "all":
                results = query.all()
            else:
                page_size = int(page_size)
                results = query.limit(page_size).offset((page - 1) * page_size).all()

            par_ids = [par.id for par in results]

            # Get latest activity status for all PARs in one query
            status_map = get_latest_par_status_map(par_ids, self._session_factory)

            # Add current_status to each PAR in the results
            for par in results:
                par.current_status = status_map.get(par.id)

            return {
                "founds": results,
                "search_options": {
                    "page": page,
                    "page_size": page_size,
                    "ordering": ordering,
                    "total_count": total_count,
                    "search_term": search_term,
                },
            }

    def track_par_view(self, par_id: str, user_id: str):
        """
        Track when a user views a PAR by updating or creating a record in the reviewed_par table.

        Args:
            par_id: The ID of the PAR being viewed
            user_id: The ID of the user viewing the PAR
        """
        with self._session_factory() as session:
            # Check if a record already exists
            reviewed_par = (
                session.query(ReviewedPar)
                .filter(ReviewedPar.par_id == par_id, ReviewedPar.user_id == user_id)
                .first()
            )

            if reviewed_par:
                # Update existing record
                reviewed_par.updated_at = datetime.utcnow()
            else:
                # Create new record
                reviewed_par = ReviewedPar(
                    par_id=par_id,
                    user_id=user_id,
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                )
                session.add(reviewed_par)

            session.commit()

    def get_list(
        self,
        schema: Any,
        searchable_fields: Optional[List[str]] = None,
        exclude_budget_info: bool = False,
    ) -> Any:

        exclude_eagers = [
            "budget_info",
            "budget_info.budget_items",
            "par_budget_analysis",
            "par_activities",
        ]

        result = self._repository.read_by_options(
            schema, searchable_fields, eager=True, exclude_eagers=exclude_eagers
        )

        # Get all PAR IDs from the results
        par_ids = [par.id for par in result.get("founds", [])]

        # Get latest activity status for all PARs in one query
        status_map = get_latest_par_status_map(par_ids, self._session_factory)

        # Add current_status to each PAR in the results
        for par in result.get("founds", []):
            par.current_status = status_map.get(par.id)

        return result
